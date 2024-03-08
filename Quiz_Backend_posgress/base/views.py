# views.py
from base64 import urlsafe_b64decode
import random
import openpyxl
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from .models import User,Question,Answer,Group,Quiz,GroupMembership
from rest_framework.permissions import IsAuthenticated
from .serializers import UserSerializer, LoginSerializer, LogoutSerializer,QuestionSerializer,AnswerSerializer,GroupSerializer,QuizSerializer,GroupMembershipSerializer,CreateQuizSerializer
from django.contrib.auth import authenticate
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes


from django.core.mail import send_mail
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.auth import get_user_model



class RegisterView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoginView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            user = authenticate(request, email=email, password=password)
            if user:
                refresh = RefreshToken.for_user(user)
                return Response({
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                })
            else:
                return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LogoutView(APIView):

    def post(self, request):
        permission_classes = [IsAuthenticated]  # Require authentication for this view

        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            try:
                refresh_token = serializer.validated_data['refresh']
                token = RefreshToken(refresh_token)
                token.blacklist()
                return Response({'message': 'Logout successful'}, status=status.HTTP_205_RESET_CONTENT)
            except Exception as e:
                return Response({'error': 'Invalid token'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


# give all question
class QuestionListView(APIView):
    def get(self,request):
        questions = list(Question.objects.all())
        random.shuffle(questions)  

        serialized_questions = QuestionSerializer(questions, many=True).data
        return Response(serialized_questions)
    

# give all quiz and a particular quiz,create a quiz(by passing id of group in which want to form quiz by admin) and delete a particular quiz
class QuizListView(APIView):
    def get(self, request, id=None):
        if id is None:
            quizzes = Quiz.objects.all()
            serializer = QuizSerializer(quizzes, many=True)
            return Response(serializer.data)
        else:
            try:
                quiz = Quiz.objects.get(id=id)
                serializer = QuizSerializer(quiz)
                return Response(serializer.data)
            except Quiz.DoesNotExist:
                return Response({"message": "Quiz not found"}, status=status.HTTP_404_NOT_FOUND)
            

    def post(self, request, id):

       # id is group_id
        group_id=id
        #  user is an admin of the group
        if not request.user.admin_of_groups.filter(id=group_id).exists():
            return Response({"message": "You are not authorized to create a quiz for this group"}, status=status.HTTP_403_FORBIDDEN)
        
        
        # Convert group_id to integer
        try:
                group_id = int(group_id)
        except ValueError:
                return Response({"message": "Invalid group ID"}, status=status.HTTP_400_BAD_REQUEST)    

        data = {
            'title': request.data.get('title'),
            'group': int(group_id ) 
            }
        print(f"after {data['group']}")
        print(data)

        serializer = CreateQuizSerializer(data=data)
        if serializer.is_valid():
            serializer.save(group_id=group_id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST) 
            
                             

            

    # permission_classes = [IsAuthenticated]

    def delete(self, request, id):
        # Check user is an admin of the group associated with the quiz
        try:
            quiz = Quiz.objects.get(id=id)
            group = quiz.group  #  ForeignKey relation between Quiz and Group
            if not request.user.admin_of_groups.filter(id=group.id).exists():
                return Response({"message": "You are not authorized to delete this quiz"}, status=status.HTTP_403_FORBIDDEN)
        except Quiz.DoesNotExist:
            return Response({"message": "Quiz not found"}, status=status.HTTP_404_NOT_FOUND)

        # Delete the quiz
        quiz.delete()
        return Response({"message": "Quiz deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        
                

# give all group and a particular group,also can create group,update group,delete group
class GroupListView(APIView):
    # permission_classes = [IsAuthenticated]  

    def get(self, request,id=None):
        if id is None:    
            groups = Group.objects.all()
            serialized_groups = GroupSerializer(groups, many=True).data
            return Response(serialized_groups)
        
        else:
            try:
                group = Group.objects.get(id=id)
                serialized_group = GroupSerializer(group).data
                return Response(serialized_group)
            except Group.DoesNotExist:
                return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)
            
    def post(self, request):
            serializer = GroupSerializer(data=request.data)
            if serializer.is_valid():
                #  admin of the group to the current user
                serializer.validated_data['admin'] = request.user
                
                # add the current user as a member
                group = serializer.save()
                group.members.add(request.user)
                group.save()
                
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, id):
        try:
            group = Group.objects.get(id=id)
            serializer = GroupSerializer(group, data=request.data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

        
    def delete(self, request,id):
        try:
            group = Group.objects.get(id=id)
            group.delete()
            return Response({"message": "Group deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)
 
               
# all quizzes that belong to group     
class QuizByGroupAPIView(APIView):
    serializer_class = QuizSerializer

    def get(self,request,group_id):
        quizs=Quiz.objects.filter(group_id=group_id)  
        serialized_quizs = QuizSerializer(quizs, many=True).data
        return Response(serialized_quizs)  

# join or remove a group
class GroupJoinRemoveView(APIView):
    serializer_class = GroupSerializer

    def post(self, request, group_id):
        try:
            group = Group.objects.get(id=group_id)
            
            if group.privacy_setting == 'public':
                # Directly add user to the group
                group.members.add(request.user)
                return Response({"message": f"Joined {group.name} successfully"}, status=status.HTTP_201_CREATED)
            else:
                membership_request = GroupMembership.objects.filter(user=request.user, group=group).first()
                if membership_request:
                    return Response({"message": "Membership request already exists"}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Create a new membership request
                    membership_request = GroupMembership.objects.create(user=request.user, group=group, status='PENDING')
                    return Response({"message": "Membership request sent for approval"}, status=status.HTTP_201_CREATED)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)    


    def delete(self, request, group_id):
        try:
            group = Group.objects.get(id=group_id)
            if request.user in group.members.all():
                group.members.remove(request.user)
                return Response({"message": f"You have left the group {group.name}"}, status=status.HTTP_204_NO_CONTENT)
            else:
                return Response({"message": "You are not a member of this group"}, status=status.HTTP_400_BAD_REQUEST)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

        


    
#  see all joining request and approve or reject that request   
class GroupMembershipRequestView(APIView):
    serializer_class = GroupMembershipSerializer

    def get(self, request, group_id):
        # current user is the admin of the group
        try:
            group = Group.objects.get(id=group_id, admin=request.user)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

        # membership requests for the group
        membership_requests = GroupMembership.objects.filter(group=group, status='PENDING')
        serializer = GroupMembershipSerializer(membership_requests, many=True)
        return Response(serializer.data)

    def put(self, request, group_id, membership_id):
        #  current user is the admin of the group
        try:
            group = Group.objects.get(id=group_id, admin=request.user)
        except Group.DoesNotExist:
            return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            membership_request = GroupMembership.objects.get(id=membership_id, group=group, status='PENDING')
            action = request.data.get('action')

            if action == 'APPROVE':
                membership_request.status = 'APPROVED'
                membership_request.save()
                group.members.add(membership_request.user)
                return Response({"message": "Membership request approved"}, status=status.HTTP_200_OK)
            elif action == 'REJECT':
                membership_request.status = 'REJECTED'
                membership_request.save()
                return Response({"message": "Membership request rejected"}, status=status.HTTP_200_OK)
            else:
                return Response({"message": "Invalid action"}, status=status.HTTP_400_BAD_REQUEST)
        except GroupMembership.DoesNotExist:
            return Response({"message": "Membership request not found"}, status=status.HTTP_404_NOT_FOUND)


# give all questions of a particular quiz in a group
class QuizinGroupQuestionView(APIView):
    # def get(self, request, group_id, quiz_id):
    def get(self, request, quiz_id):

        # questions = Question.objects.filter(quiz__group_id=group_id, quiz_id=quiz_id)
        questions = list(Question.objects.filter(quiz_id=quiz_id))

        random.shuffle(questions)  


        # Serialize the questions and their answers
        serialized_questions = QuestionSerializer(questions, many=True).data
        return Response(serialized_questions)


    



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_questions_from_excel(request, quiz_id):
    admin = request.user

    # Check if the user is an admin of the group associated with the quiz
    quiz = get_object_or_404(Quiz, id=quiz_id)
    group = quiz.group
    if not (admin.is_superuser or admin == group.admin):
        return Response({"error": "Only admin of the group can import questions and answers."}, status=403)
    


    # file_path = request.data.get('file_path')  # Ensure you pass the file path in the request data
    # if not file_path:
    #     return Response({"error": "File path is required."}, status=400)

    workbook = openpyxl.load_workbook('base/questionsandanswers.xlsx')
    worksheet = workbook.active

    for row in worksheet.iter_rows(values_only=True):
        question_text = row[0]
        options = row[1:5]
        correct_index = row[5]
        if question_text is None:
            break

        print(f"question_text:{question_text} options:{options} correct_index:{correct_index}")

        # Create the question
        question = Question.objects.create(question=question_text, quiz_id=quiz_id)

        # Create answers
        for index, option in enumerate(options):
            is_correct = True if index == correct_index else False
            Answer.objects.create(question=question, answer=option, is_correct=is_correct)

    return Response({"message": "Questions and answers imported successfully."}, status=201)






@api_view(['POST'])
def create_group_and_import_questions(request, group_id):
    admin = request.user

    # Extract quiz details 
    name = request.data.get('name',"Default Quiz Name" )

    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return Response({"message": "Group not found"}, status=status.HTTP_404_NOT_FOUND)

    # Create the quiz within the existing group
    quiz_data = {'title': name, 'group': group_id}
    serializer_quiz = QuizSerializer(data=quiz_data)
    if serializer_quiz.is_valid():
        quiz = serializer_quiz.save()
    else:
        return Response(serializer_quiz.errors, status=status.HTTP_400_BAD_REQUEST)

    if not admin.admin_of_groups.filter(id=group_id).exists():
        quiz.delete()  # Rollback quiz creation if user is not an admin
        return Response({"message": "You are not authorized to create a quiz for this group"}, status=status.HTTP_403_FORBIDDEN)

    try:
        workbook = openpyxl.load_workbook('base/questionsandanswers.xlsx')
        worksheet = workbook.active

        for row in worksheet.iter_rows(values_only=True):
            question_text = row[0]
            options = row[1:5]
            correct_index = row[5]
            if question_text is None:
                break

            #  question
            question = Question.objects.create(question=question_text, quiz=quiz)

            #  answers
            for index, option in enumerate(options):
                is_correct = True if index == correct_index else False
                Answer.objects.create(question=question, answer=option, is_correct=is_correct)
    except Exception as e:
        # Rollback quiz creation if an error occurs during question import
        quiz.delete()
        return Response({"message": "Failed to import questions. Error: {}".format(str(e))}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({"message": "Quiz created successfully."}, status=status.HTTP_201_CREATED)










class ForgotPasswordView(APIView):
    def post(self, request):
        email = request.data.get('email')
        try:
            user = User.objects.get(email=email)
            # Generate and save a password reset token for the user
            token = default_token_generator.make_token(user)
            user.reset_password_token = token
            user.save()
            # Send password reset email
            reset_link = f"http://localhost:8000/{urlsafe_base64_encode(force_bytes(user.pk))}/{token}/"
            email_subject = 'Password Reset'
            email_body = f'Please click the following link to reset your password: {reset_link}'
            send_mail(email_subject, email_body, 'from@example.com', [email])
            return Response({'message': 'Password reset instructions sent to your email'}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'User with this email does not exist'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ResetPasswordView(APIView):
    def post(self, request, uidb64, token):
        try:
            uid = str(urlsafe_b64decode(uidb64), 'utf-8')

            user = User.objects.get(pk=uid)
            if default_token_generator.check_token(user, token):
                new_password = request.data.get('new_password')
                user.set_password(new_password)
                user.reset_password_token = None  # Clear the reset token
                user.save()
                return Response({'message': 'Password reset successful'}, status=status.HTTP_200_OK)
            else:
                return Response({'error': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)